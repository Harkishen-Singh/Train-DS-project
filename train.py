import csv
import random
from openpyxl import Workbook

'''csv_file = open('./datasets/train.csv', 'r')
reader = csv.reader(csv_file)
for row in reader:
    print(row)'''

class train:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Record_of_training"



        self.survival_final = 0 # head of all variables

        self.id = []
        self.survived = []
        self.pclass = []
        self.name = []
        self.sex = []
        self.age = []
        self.sibsp = []
        self.parch = []
        self.ticket = []
        self.fare = []
        self.cabin = []
        self.embarked = []
        self.no_of_rows = -1
        '''for i in self.reader:
            self.no_of_rows = self.no_of_rows + 1
        '''
        print(self.no_of_rows)


        self.male_class = [3]
        self.male_class_influence = [3]
        self.male_class_count = [3]
        self.male_class_count[0] = 0
        self.male_class_count[1] = 0
        self.male_class_count[2] = 0

        self.female_class = [3]
        self.female_class_influence = [3]
        self.female_class_count = [3]
        self.female_class_count[0] = 0
        self.female_class_count[1] = 0
        self.female_class_count[2] = 0

        # independent of gender
        self.cabin_count = 0
        self.cabin_count_alpha = [6]  # A, B, C, D ,E ,F
        self.cabin_alpha_influence = [6]
        self.cabin_count_alpha_survived = [6]

        # variable for embarked place
        self.embarked_place_count = [3] # S , C , Q
        self.embarked_place_count_survived = [3]
        self.embarked_influence = [3]

    def asking_values(self):
        self.fname = input('Enter .csv filename to train the machine : ')
        ifile = open('./datasets/'+self.fname, 'r')
        self.reader = csv.reader(ifile)

        '''self.csv_file = open('./datasets/train.csv', 'r')
        self.reader = csv.reader(self.csv_file)'''

        for i in self.reader:
            self.id.append(str(i[0]))
            self.survived.append(i[1])
            self.pclass.append(i[2])
            self.name.append(i[3])
            self.sex.append(i[4])
            self.age.append(i[5])
            self.sibsp.append(i[6])
            self.parch.append(i[7])
            self.ticket.append(i[8])
            self.fare.append(i[9])
            self.cabin.append(i[10])
            self.embarked.append(i[11])
        self.count = len(self.id)

    def assigningValues(self):
        print('got in')

        print(len(self.id))
        '''for i in range(1, len(self.id)):
            print(self.id[i] + ' ' + self.survived[i] + ' ' + self.pclass[i] + ' ' + self.name[i])'''

    def calculating_influence(self):

        # initialisation for various parameters

        for i in range(0,3):
            self.embarked_influence[i] = 0
        for i in range(0,3):
            self.embarked_place_count[i] = 0
        for i in range(0,3):
            self.embarked_place_count_survived[i] = 0

        for i in range(0,6):
            self.cabin_alpha_influence[i] = 0
        for i in range(0,6):
            self.cabin_count_alpha[i] = 0
        for i in range(0,6):
            self.cabin_count_alpha_survived[i] = 0




        for i in range(1, self.count):
            if self.sex[i]=='male' :

                # male code below

                # below is ML of class vs survivial

                if self.pclass[i] == '1':
                     if self.survived[i] == '1' :
                        self.male_class[0] = self.male_class[0] + 1
                     self.male_class_count[0] = self.male_class_count[0] + 1

                if self.pclass[i] == '2':
                     if self.survived[i] == '1':
                        self.male_class[1] = self.male_class[1] + 1
                     self.male_class_count[1] = self.male_class_count[1] + 1

                if self.pclass[i] == '3':
                     if self.survived[i] == '1':
                        self.male_class[2] = self.male_class[2] + 1
                     self.male_class_count[2] = self.male_class_count[2] + 1

                # saving the probability of survival in case of men, class wise
                ''' write this code after the loop of i ends
                self.male_class_influence[0] = self.male_class[0] / self.male_class_count[0] # class 1
                self.male_class_influence[1] = self.male_class[1] / self.male_class_count[1] # class 2
                self.male_class_influence[2] = self.male_class[2] / self.male_class_count[2] # class 3
                '''

            if self.sex[i]=='female' :

                # female code below

                # below is ML of class vs survivial

                if self.pclass[i] == '1':
                     if self.survived[i] == '1' :
                        self.female_class[0] = self.female_class[0] + 1
                     self.female_class_count[0] = self.female_class_count[0] + 1

                if self.pclass[i] == '2':
                     if self.survived[i] == '1':
                        self.female_class[1] = self.female_class[1] + 1
                     self.female_class_count[1] = self.female_class_count[1] + 1

                if self.pclass[i] == '3':
                     if self.survived[i] == '1':
                        self.female_class[2] = self.female_class[2] + 1
                     self.female_class_count[2] = self.female_class_count[2] + 1

                # saving the probability of survival in case of men, class wise
                '''
                self.female_class_influence[0] = self.female_class[0] / self.female_class_count[0] # class 1
                self.female_class_influence[1] = self.female_class[1] / self.female_class_count[1] # class 2
                self.female_class_influence[2] = self.female_class[2] / self.female_class_count[2] # class 3
                '''
                # influence means the probability of survival based on class

            # cabin probability
            if self.cabin == '':
                a = 0 # just as a null statement, ignore it
            else:
                self.cabin_count = self.cabin_count + 1

                if self.cabin[0:1] == 'A' :
                    self.cabin_count_alpha[0] = self.cabin_count_alpha[0] + 1
                    if self.survived == '1':
                        self.cabin_count_alpha_survived[0] = self.cabin_count_alpha_survived[0] + 1

                if self.cabin[0:1] == 'B' :
                    self.cabin_count_alpha[1] = self.cabin_count_alpha[1] + 1
                    if self.survived == '1':
                        self.cabin_count_alpha_survived[1] = self.cabin_count_alpha_survived[1] + 1

                if self.cabin[0:1] == 'C' :
                    self.cabin_count_alpha[2] = self.cabin_count_alpha[2] + 1
                    if self.survived == '1':
                        self.cabin_count_alpha_survived[2] = self.cabin_count_alpha_survived[2] + 1

                if self.cabin[0:1] == 'D' :
                    self.cabin_count_alpha[3] = self.cabin_count_alpha[3] + 1
                    if self.survived == '1':
                        self.cabin_count_alpha_survived[3] = self.cabin_count_alpha_survived[3] + 1

                if self.cabin[0:1] == 'E' :
                    self.cabin_count_alpha[4] = self.cabin_count_alpha[4] + 1
                    if self.survived == '1':
                        self.cabin_count_alpha_survived[4] = self.cabin_count_alpha_survived[4] + 1

                if self.cabin[0:1] == 'F' :
                    self.cabin_count_alpha[5] = self.cabin_count_alpha[5] + 1
                    if self.survived == '1':
                        self.cabin_count_alpha_survived[5] = self.cabin_count_alpha_survived[5] + 1

                '''write this code after the loop of i ends
                self.cabin_alpha_influence[0] = self.cabin_count_alpha_survived[0] / self.cabin_count_alpha[0]
                self.cabin_alpha_influence[1] = self.cabin_count_alpha_survived[1] / self.cabin_count_alpha[1]
                self.cabin_alpha_influence[2] = self.cabin_count_alpha_survived[2] / self.cabin_count_alpha[2]
                self.cabin_alpha_influence[3] = self.cabin_count_alpha_survived[3] / self.cabin_count_alpha[3]
                self.cabin_alpha_influence[4] = self.cabin_count_alpha_survived[4] / self.cabin_count_alpha[4]
                self.cabin_alpha_influence[5] = self.cabin_count_alpha_survived[5] / self.cabin_count_alpha[5]
                '''

            # influence of place embarked below

            if self.embarked[i] == 'S' :
                if self.survived == '1' :
                    self.embarked_place_count_survived[0] = self.embarked_place_count_survived[0] +1
                self.embarked_place_count[0] = self.embarked_place_count[0] +1
            if self.embarked[i] == 'C' :
                if self.survived == '1' :
                    self.embarked_place_count_survived[1] = self.embarked_place_count_survived[1] +1
                self.embarked_place_count[1] = self.embarked_place_count[1] +1
            if self.embarked[i] == 'Q' :
                if self.survived == '1' :
                    self.embarked_place_count_survived[0] = self.embarked_place_count_survived[0] +1
                self.embarked_place_count[1] = self.embarked_place_count[1] +1

            '''write this code after loop i completes
            self.embarked_influence[0] = self.embarked_place_count_survived[0] / self.embarked_place_count[0] # S
            self.embarked_influence[1] = self.embarked_place_count_survived[1] / self.embarked_place_count[1] # C
            self.embarked_influence[2] = self.embarked_place_count_survived[2] / self.embarked_place_count[2] # Q
            '''


        # i loop ends here

        self.male_class_influence[0] = self.male_class[0] / self.male_class_count[0]  # class 1
        self.male_class_influence[1] = self.male_class[1] / self.male_class_count[1]  # class 2
        self.male_class_influence[2] = self.male_class[2] / self.male_class_count[2]  # class 3

        self.female_class_influence[0] = self.female_class[0] / self.female_class_count[0]  # class 1
        self.female_class_influence[1] = self.female_class[1] / self.female_class_count[1]  # class 2
        self.female_class_influence[2] = self.female_class[2] / self.female_class_count[2]  # class 3

        self.cabin_alpha_influence[0] = self.cabin_count_alpha_survived[0] / self.cabin_count_alpha[0]
        self.cabin_alpha_influence[1] = self.cabin_count_alpha_survived[1] / self.cabin_count_alpha[1]
        self.cabin_alpha_influence[2] = self.cabin_count_alpha_survived[2] / self.cabin_count_alpha[2]
        self.cabin_alpha_influence[3] = self.cabin_count_alpha_survived[3] / self.cabin_count_alpha[3]
        self.cabin_alpha_influence[4] = self.cabin_count_alpha_survived[4] / self.cabin_count_alpha[4]
        self.cabin_alpha_influence[5] = self.cabin_count_alpha_survived[5] / self.cabin_count_alpha[5]

        self.embarked_influence[0] = self.embarked_place_count_survived[0] / self.embarked_place_count[0]  # S
        self.embarked_influence[1] = self.embarked_place_count_survived[1] / self.embarked_place_count[1]  # C
        self.embarked_influence[2] = self.embarked_place_count_survived[2] / self.embarked_place_count[2]  # Q

    def saving_data_to_excel(self):
        # naming headings
        self.ws.cell(row=1, column=1, value='male_class_influence')
        self.ws.cell(row=1, column=2, value='female_class_influence')
        self.ws.cell(row=1, column=3, value='cabin_alpha_influence')
        self.ws.cell(row=1, column=4, value='embarked_influence')

        for i in range(1,4):
            self.ws.cell(row=i+1, column=1, value=self.male_class_influence)
            self.ws.cell(row=i+1, column=2, value=self.female_class_influence)
            self.ws.cell(row=i+1, column=4, value=self.embarked_influence)

        for i in range(2,7):
            self.ws.cell(row=i, column=3, value=self.cabin_alpha_influence)

        self.wb.save(self.fname+'Processed.xlsm')


obj = train()
obj.assigningValues()