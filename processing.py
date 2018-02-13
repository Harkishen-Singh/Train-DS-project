from openpyxl import load_workbook
import csv

'''This file analyses the input file based on learnt records, learned during machine learning part exeution'''

class Processing():

    def __init__(self):

        self.wb = load_workbook('./datasets/savedProcessed/trainProcessed.xlsm')
        self.ws = self.wb['Record_of_training']
        self.inf_male_class = [0, 0, 0]
        self.inf_female_class = [0, 0, 0]
        self.inf_cabin = [0, 0, 0, 0, 0, 0]
        self.inf_embarked = [0, 0, 0]
        for i in range(0, 3):
            self.inf_male_class[i] = self.ws.cell(row=i+2,column=1).value
            self.inf_female_class[i] = self.ws.cell(row=i + 2, column=2).value
            self.inf_embarked[i] = self.ws.cell(row = i+2, column=4).value
        for i in range(0,6):
            self.inf_cabin[i] = self.ws.cell(row = i+2, column = 3).value

        self.parameters = []

    def input_file_storage(self):
        fname = input('.csv File name to be predicted :')
        self.reader = csv.reader(fname)
        self.name = []
        self.id = []
        self.pclass = []        # dont forget to make survival for age > 50 as 0 in male
        self.sex = []
        self.name = []
        self.embark = []
        self.cabin = []
        self.age = []
        self.survival = 0  # main head of all variables
        self.cal_sur = [] # calculating survival
        for i in self.reader:
            self.id.append(i[0])
            self.name.append(i[2])
            self.pclass.append(i[1])
            self.sex.append(i[3])
            self.age.append(i[4])
            self.cabin.append(i[9])
            self.embark.append(i[10])
            self.parameters.append(0)
            self.cal_sur.append(0)
        self.count = len(self.id)

    def analysis(self):

        for i in range(0, self.count):

            if self.sex[i] == 'male':
                if self.pclass[i] == '1':
                    self.parameters[i] = self.parameters[i] +1
                    self.cal_sur[i] = self.inf_male_class[0] + self.cal_sur[i]

                if self.pclass[i] == '2':
                    self.parameters[i] = self.parameters[i] +1
                    self.cal_sur[i] = self.inf_male_class[1]+ self.cal_sur[i]

                if self.pclass[i] == '3':
                    self.parameters[i] = self.parameters[i] +1
                    self.cal_sur[i] = self.inf_male_class[2]+ self.cal_sur[i]

            if self.sex[i] == 'female':
                if self.pclass[i] == '1':
                    self.parameters[i] = self.parameters[i] + 1
                    self.cal_sur[i] = self.inf_female_class[0]+ self.cal_sur[i]

                if self.pclass[i] == '2':
                    self.parameters[i] = self.parameters[i] + 1
                    self.cal_sur[i] = self.inf_female_class[1]+ self.cal_sur[i]

                if self.pclass[i] == '3':
                    self.parameters[i] = self.parameters[i] + 1
                    self.cal_sur[i] = self.inf_female_class[2]+ self.cal_sur[i]
            if self.cabin[i] == '':
                a=0
            else:
                if self.cabin[i][0:1] == 'A':
                    self.parameters[i] = self.parameters[i] +1
                    self.cal_sur[i] =  self.inf_cabin[0] + self.cal_sur[i]

                if self.cabin[i][0:1] == 'B':
                    self.parameters[i] = self.parameters[i] +1
                    self.cal_sur[i] =  self.inf_cabin[1] + self.cal_sur[i]
                if self.cabin[i][0:1] == 'C':
                    self.parameters[i] = self.parameters[i] +1
                    self.cal_sur[i] =  self.inf_cabin[2] + self.cal_sur[i]
                if self.cabin[i][0:1] == 'D':
                    self.parameters[i] = self.parameters[i] +1
                    self.cal_sur[i] =  self.inf_cabin[3] + self.cal_sur[i]
                if self.cabin[i][0:1] == 'E':
                    self.parameters[i] = self.parameters[i] +1
                    self.cal_sur[i] =  self.inf_cabin[4] + self.cal_sur[i]
                if self.cabin[i][0:1] == 'F':
                    self.parameters[i] = self.parameters[i] +1
                    self.cal_sur[i] =  self.inf_cabin[5] + self.cal_sur[i]

            if self.embark[i] == 'S':
                self.parameters[i] = self.parameters[i] +1
                self.cal_sur[i] = self.cal_sur[i] + self.inf_embarked[0]
            if self.embark[i] == 'C':
                self.parameters[i] = self.parameters[i] +1
                self.cal_sur[i] = self.cal_sur[i] + self.inf_embarked[1]
            if self.embark[i] == 'Q':
                self.parameters[i] = self.parameters[i] +1
                self.cal_sur[i] = self.cal_sur[i] + self.inf_embarked[2]

        # end of loop i here

        


